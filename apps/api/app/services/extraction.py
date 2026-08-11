"""Document → structured invoice extraction.

Pulls text from PDFs/Word/Excel/CSV and hands the document to the AI provider for
structured extraction. Scanned PDFs and images are passed as raw bytes so the
provider can OCR them natively (Claude document/vision input). Parsing libraries are
optional imports — the pipeline degrades to AI-only extraction if a parser is absent.

Returns a list of (source_label, Invoice) so a single upload (or ZIP) can yield
multiple invoices.
"""

from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass, field

from ..ai.base import ExtractionInput
from ..ai.factory import get_ai_provider
from ..vat.schemas import Invoice
from . import ocr

_IMAGE_EXT = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".tiff", ".bmp")
_SUPPORTED_EXT = (".pdf", ".docx", ".doc", ".xlsx", ".csv", ".txt", ".zip", *_IMAGE_EXT)


class UnsupportedFileError(ValueError):
    """Raised when an uploaded file's type is not supported."""


@dataclass
class ExtractedDoc:
    label: str
    invoice: Invoice
    raw_text: str = ""
    ocr_used: bool = False
    ocr_engine: str | None = None
    page_count: int = 0
    warnings: list[str] = field(default_factory=list)


def _pdf_text(data: bytes) -> str:
    return ocr.extract_pdf_text(data).text


def _docx_text(data: bytes) -> str:
    try:
        import docx

        document = docx.Document(io.BytesIO(data))
        return "\n".join(p.text for p in document.paragraphs).strip()
    except Exception:  # noqa: BLE001
        return ""


def _doc_text(data: bytes) -> str:
    """Best-effort text extraction from a legacy binary Word (.doc) file.

    There is no pure-Python parser for the old OLE `.doc` format in the base
    dependencies, so we isolate the `WordDocument` stream (via ``olefile`` when
    available) and scrape readable UTF-16LE / ASCII runs. Good enough to give the
    assistant the document's text; the caller warns if the result looks empty.
    """
    import re

    payload = data
    try:
        import olefile  # optional; cleaner extraction when present

        bio = io.BytesIO(data)
        if olefile.isOleFile(bio):
            ole = olefile.OleFileIO(bio)
            if ole.exists("WordDocument"):
                payload = ole.openstream("WordDocument").read()
            ole.close()
    except Exception:  # noqa: BLE001
        pass

    # UTF-16LE runs cover most modern .doc text; ASCII runs catch older content.
    utf16 = re.findall(rb"(?:[\x20-\x7e]\x00){4,}", payload)
    u16 = " ".join(seg.decode("utf-16-le", "ignore") for seg in utf16).strip()
    ascii_runs = re.findall(rb"[\x20-\x7e]{4,}", payload)
    asc = " ".join(seg.decode("latin-1", "ignore") for seg in ascii_runs).strip()
    text = u16 if len(u16) >= len(asc) else asc
    return re.sub(r"[ \t]{2,}", " ", text).strip()


def _xlsx_text(data: bytes) -> str:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                if any(cells):
                    lines.append("\t".join(cells))
        return "\n".join(lines).strip()
    except Exception:  # noqa: BLE001
        return ""


def _csv_text(data: bytes) -> str:
    try:
        text = data.decode("utf-8", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        return "\n".join("\t".join(r) for r in rows).strip()
    except Exception:  # noqa: BLE001
        return ""


def _read_text(filename: str, data: bytes) -> ocr.TextExtraction:
    """Get raw text (with OCR where needed) and extraction metadata for a file."""
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return ocr.extract_pdf_text(data)
    if lower.endswith(_IMAGE_EXT):
        return ocr.extract_image_text(data)
    if lower.endswith(".docx"):
        return ocr.TextExtraction(text=_docx_text(data), page_count=1)
    if lower.endswith(".doc"):
        return ocr.TextExtraction(text=_doc_text(data), page_count=1)
    if lower.endswith(".xlsx"):
        return ocr.TextExtraction(text=_xlsx_text(data), page_count=1)
    if lower.endswith(".csv"):
        return ocr.TextExtraction(text=_csv_text(data), page_count=1)
    if lower.endswith(".txt"):
        return ocr.TextExtraction(text=data.decode("utf-8", errors="replace"), page_count=1)
    raise UnsupportedFileError(f"Unsupported file type: {filename}")


def _build_input(filename: str, data: bytes, mime: str | None, text: str | None) -> ExtractionInput:
    # Raw document bytes are attached for PDFs/images so an AI provider can use native
    # document/vision input; other formats pass their extracted text only.
    lower = filename.lower()
    doc_bytes = data if lower.endswith((".pdf", *_IMAGE_EXT)) else None
    return ExtractionInput(filename=filename, mime=mime, text=text or None, doc_bytes=doc_bytes)


def _norm_txt(s: str) -> str:
    import re
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _attach_bboxes(invoice: Invoice, layout: list) -> None:
    """Give each extracted field's evidence a page + normalised bbox by matching its source
    snippet to a layout line — so the UI can highlight the region on the page image."""
    if not invoice.field_evidence or not layout:
        return
    lines = [(_norm_txt(ln.text), ln) for ln in layout]
    for ev in invoice.field_evidence.values():
        snip = _norm_txt(ev.get("snippet", ""))
        if not snip:
            continue
        match = None
        for nt, ln in lines:
            if nt and (snip in nt or nt in snip):
                match = ln
                break
        if match:
            ev["page"] = match.page
            ev["bbox"] = [round(c, 4) for c in match.bbox]


def _extract_one(filename: str, data: bytes, mime: str | None) -> ExtractedDoc:
    te = _read_text(filename, data)
    provider = get_ai_provider()
    source = _build_input(filename, data, mime, te.text)
    invoice = provider.extract_invoice(source)
    try:
        _attach_bboxes(invoice, ocr.extract_layout(filename, data))
    except Exception:  # noqa: BLE001 — highlighting is a nicety, never break extraction
        pass
    return ExtractedDoc(
        label=filename,
        invoice=invoice,
        raw_text=te.text,
        ocr_used=te.ocr_used,
        ocr_engine=te.engine,
        page_count=te.page_count,
        warnings=list(te.warnings),
    )


def document_text(filename: str, data: bytes) -> str:
    """Extract plain text from a document for knowledge-base ingestion."""
    lower = filename.lower()
    if lower.endswith(".pdf"):
        return _pdf_text(data)
    if lower.endswith(".docx"):
        return _docx_text(data)
    if lower.endswith(".xlsx"):
        return _xlsx_text(data)
    if lower.endswith(".csv"):
        return _csv_text(data)
    return data.decode("utf-8", errors="replace")


def extract_invoices(filename: str, data: bytes, mime: str | None = None) -> list[ExtractedDoc]:
    """Extract one or more invoices from an uploaded file (ZIP -> many)."""
    if filename.lower().endswith(".zip"):
        results: list[ExtractedDoc] = []
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist():
                if info.is_dir() or info.filename.startswith("__MACOSX"):
                    continue
                if not info.filename.lower().endswith(_SUPPORTED_EXT):
                    continue  # skip unsupported entries silently within a batch
                inner = zf.read(info)
                try:
                    results.append(_extract_one(info.filename, inner, None))
                except UnsupportedFileError:
                    continue
        if not results:
            raise UnsupportedFileError("ZIP contained no supported documents.")
        return results

    return [_extract_one(filename, data, mime)]
