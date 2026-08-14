"""Parse a CSV/Excel transactions file into normalised Transactions.

Column mapping is automatic and layout-agnostic: headers are matched to fields by
keyword, and direction/treatment/emirate are inferred from values when not explicit.
No per-file template.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from .schemas import Direction, Emirate, Transaction, Treatment

# Field → candidate header keywords, in assignment priority order (specific first).
_FIELD_ORDER: list[tuple[str, list[str]]] = [
    ("date", ["invoicedate", "taxdate", "documentdate", "postingdate", "txndate", "date"]),
    ("trn", ["trn", "taxregistrationnumber", "vatnumber", "vatno", "taxregistration"]),
    ("invoice_number", ["invoiceno", "invoicenumber", "billno", "documentno", "docno", "reference", "refno", "invoice"]),
    ("emirate", ["emirate", "region"]),
    ("vat_rate", ["vatrate", "taxrate", "rate"]),
    ("treatment", ["taxcode", "vatcode", "vattype", "taxtype", "treatment", "code"]),
    ("vat_amount", ["vatamount", "taxamount", "outputvat", "inputvat", "vat", "tax"]),
    ("taxable_amount", ["taxableamount", "taxablevalue", "netamount", "subtotal", "amountexclvat", "net", "amount", "value"]),
    ("doc_type", ["documenttype", "transactiontype", "doctype", "type"]),
    ("direction", ["direction", "salespurchase", "inout"]),
    ("party", ["customername", "suppliername", "vendor", "customer", "supplier", "client", "party", "name", "account"]),
]

_DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
    "%d %b %Y", "%d %B %Y", "%b %d %Y", "%B %d %Y", "%b %d, %Y", "%B %d, %Y", "%d-%b-%Y",
]

_EMIRATES = {
    "abudhabi": Emirate.ABU_DHABI, "auh": Emirate.ABU_DHABI,
    "dubai": Emirate.DUBAI, "dxb": Emirate.DUBAI,
    "sharjah": Emirate.SHARJAH, "shj": Emirate.SHARJAH,
    "ajman": Emirate.AJMAN,
    "ummalquwain": Emirate.UMM_AL_QUWAIN, "uaq": Emirate.UMM_AL_QUWAIN,
    "rasalkhaimah": Emirate.RAS_AL_KHAIMAH, "rak": Emirate.RAS_AL_KHAIMAH,
    "fujairah": Emirate.FUJAIRAH, "fuj": Emirate.FUJAIRAH,
}


def _norm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def _clean_id(v) -> str | None:
    """Return an identifier (TRN / invoice number) as its full text. Excel stores long
    digit strings as floats and shows them in scientific notation (e.g. a 15-digit TRN
    becomes '1.00047E+14'); recover the full integer digits so the TRN isn't corrupted."""
    if v is None:
        return None
    if isinstance(v, float):
        return format(int(v), "d") if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    if not s:
        return None
    # A scientific-notation string like "1.00047E+14" or "1,00047E+14".
    if re.fullmatch(r"[+-]?\d[\d,]*\.?\d*\s*[eE]\s*[+-]?\d+", s):
        try:
            return format(int(float(s.replace(",", ""))), "d")
        except (ValueError, OverflowError):
            return s
    return s


def _to_decimal(v) -> Decimal:
    if v is None:
        return Decimal(0)
    s = str(v).strip().replace(",", "").replace("AED", "").replace("(", "-").replace(")", "")
    if s in ("", "-", "--"):
        return Decimal(0)
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal(0)


def parse_date(v) -> date | None:
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


# Normalise transaction-type values (incl. the FTA FAF codes "01 Invoice",
# "02 Credit Note", "03 Debit Note") to clean labels for the drill-down.
_DOC_TYPE_MAP: list[tuple[tuple[str, ...], str]] = [
    (("creditnote", "creditmemo", "salesreturn", "salesreturns"), "Credit Note"),
    (("debitnote",), "Debit Note"),
    (("taxinvoice", "invoice"), "Invoice"),
    (("receipt",), "Receipt"),
    (("advancepayment", "advance"), "Advance Payment"),
    (("retention",), "Retention"),
    (("bill",), "Bill"),
    (("expense",), "Expense"),
    (("payment",), "Payment"),
]


def _normalize_doc_type(raw) -> str | None:
    if raw in (None, ""):
        return None
    n = _norm(str(raw))
    for keys, label in _DOC_TYPE_MAP:
        if any(k in n for k in keys):
            return label
    return str(raw).strip() or None


def _map_columns(headers: list[str]) -> dict[str, str]:
    norm = {h: _norm(h) for h in headers}
    used: set[str] = set()
    mapping: dict[str, str] = {}
    for field, keywords in _FIELD_ORDER:
        chosen = None
        for kw in keywords:
            for h in headers:
                if h in used:
                    continue
                if norm[h] == kw or kw in norm[h]:
                    chosen = h
                    break
            if chosen:
                break
        if chosen:
            mapping[field] = chosen
            used.add(chosen)
    return mapping


def _infer_direction(doc_type: str | None, direction_val: str | None) -> Direction | None:
    txt = _norm(f"{direction_val or ''} {doc_type or ''}")
    if any(k in txt for k in ("purchase", "bill", "expense", "vendor", "supplier", "input", "debit", "import")):
        return Direction.PURCHASE
    if any(k in txt for k in ("sale", "invoice", "customer", "output", "credit")):
        return Direction.SALE
    return None


def _infer_treatment(code: str | None, rate: Decimal | None, vat: Decimal, doc_type: str | None) -> Treatment | None:
    c = _norm(f"{code or ''} {doc_type or ''}")
    if any(k in c for k in ("reversecharge", "rcm", "reverse")) or c.startswith("rc"):
        return Treatment.REVERSE_CHARGE
    if "import" in c or c.startswith("im"):
        return Treatment.IMPORT
    if "exempt" in c or c.startswith("ex"):
        return Treatment.EXEMPT
    if "zero" in c or c.startswith("zr") or c.startswith("z"):
        return Treatment.ZERO_RATED
    if "standard" in c or c.startswith("sr") or c.startswith("s5"):
        return Treatment.STANDARD
    if "outofscope" in c or c.startswith("os"):
        return Treatment.OUT_OF_SCOPE
    if rate is not None:
        if rate >= Decimal("0.04"):
            return Treatment.STANDARD
        if rate == 0 and vat == 0:
            return Treatment.ZERO_RATED
    if vat > 0:
        return Treatment.STANDARD
    return None


def _rate(v) -> Decimal | None:
    s = str(v or "").strip().replace("%", "")
    if not s:
        return None
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return d / 100 if d > 1 else d


def _sheet_defaults(sheet_name: str) -> tuple[Direction | None, Treatment | None]:
    """Infer the direction + treatment implied by a sheet's name (e.g. a workbook
    with 'Standard-rated Sales' / 'Zero-rated Sales' / 'Taxable Expenses' tabs)."""
    n = _norm(sheet_name)
    direction = None
    if any(k in n for k in ("expense", "purchase", "input", "payable", "supplier", "bill")):
        direction = Direction.PURCHASE
    elif any(k in n for k in ("sale", "supply", "supplies", "output", "revenue", "income", "customer")):
        direction = Direction.SALE
    treatment = None
    if "zero" in n:
        treatment = Treatment.ZERO_RATED
    elif "exempt" in n:
        treatment = Treatment.EXEMPT
    elif "reverse" in n or "rcm" in n:
        treatment = Treatment.REVERSE_CHARGE
    elif "import" in n:
        treatment = Treatment.IMPORT
    elif "standard" in n or "taxable" in n or "5" in n:
        treatment = Treatment.STANDARD
    return direction, treatment


def _sheets_from_bytes(filename: str, data: bytes) -> list[tuple[str, list[dict]]]:
    """Return [(sheet_name, rows)]. Excel yields every sheet; CSV yields one.

    Format is detected from the CONTENT, not just the extension, so a file that is
    named .xlsx but is really a CSV (a common export) still parses, and an unsupported
    old .xls (OLE2) gives a clear message instead of an uncaught error.
    """
    is_ooxml = data[:2] == b"PK"                       # real .xlsx/.xlsm (zip container)
    is_ole = data[:4] == b"\xd0\xcf\x11\xe0"            # legacy .xls (OLE2 compound file)
    wants_excel = filename.lower().endswith((".xlsx", ".xlsm"))

    if (wants_excel or is_ooxml) and is_ooxml:
        from openpyxl import load_workbook

        try:
            wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise ValueError(
                "The Excel file could not be read. Re-save it as .xlsx (Excel Workbook) "
                "or export it as .csv and upload again."
            ) from exc
        out: list[tuple[str, list[dict]]] = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(c) if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
            out.append((ws.title, [dict(zip(headers, r)) for r in rows[1:]]))
        return out

    if is_ole or filename.lower().endswith(".xls"):
        raise ValueError(
            "The old .xls format is not supported. Open the file in Excel and save it as "
            ".xlsx (Excel Workbook) or .csv, then upload again."
        )

    # Everything else (real .csv, or a .xlsx that is actually CSV text) → parse as CSV.
    if b"\x00" in data[:4096]:  # NUL bytes ⇒ binary, not text/CSV
        raise ValueError(
            "This file isn't a readable CSV or Excel workbook. Open it in Excel and use "
            "File → Save As → Excel Workbook (.xlsx) or CSV (.csv), then upload again."
        )
    text = data.decode("utf-8-sig", errors="replace")
    # A high ratio of replacement characters also means it wasn't really text.
    if text and text.count("�") > max(20, len(text) // 20):
        raise ValueError(
            "This file isn't a readable CSV or Excel workbook. Re-save it as .xlsx or .csv "
            "and upload again."
        )
    return [("", list(csv.DictReader(io.StringIO(text))))]


def parse_emirate(value) -> Emirate | None:
    """Parse an emirate name/abbreviation (e.g. 'Dubai', 'DXB', 'abu_dhabi')."""
    if not value:
        return None
    key = _norm(str(value))
    if key in _EMIRATES:
        return _EMIRATES[key]
    for e in Emirate:
        if e.value == key or _norm(e.value) == key:
            return e
    return None


def parse_transactions(
    filename: str,
    data: bytes,
    period_start: date | None = None,
    period_end: date | None = None,
    default_emirate: Emirate | None = None,
) -> tuple[list[Transaction], dict[str, str]]:
    """Parse a CSV/Excel file into normalised transactions. Multi-sheet Excel is
    supported: each sheet's name supplies the default direction/treatment for its
    rows (overridden by explicit per-row columns where present). Date filtering only
    applies when a period range is supplied."""
    sheets = _sheets_from_bytes(filename, data)
    txns: list[Transaction] = []
    last_map: dict[str, str] = {}
    all_headers: list[str] = []
    total_rows = 0

    for sheet_name, rows in sheets:
        if not rows:
            continue
        total_rows += len(rows)
        all_headers = list(rows[0].keys())
        m = _map_columns(all_headers)
        last_map = m or last_map
        sheet_dir, sheet_treat = _sheet_defaults(sheet_name)
        # When the file identifies transactions by invoice/party, a row missing BOTH is
        # a totals/empty row — exclude it so column totals aren't double-counted.
        has_id_cols = ("invoice_number" in m) or ("party" in m)

        def cell(row, field, _m=m):
            col = _m.get(field)
            return row.get(col) if col else None

        for i, row in enumerate(rows):
            d = parse_date(cell(row, "date"))
            if period_start and period_end and d and not (period_start <= d <= period_end):
                continue
            rate = _rate(cell(row, "vat_rate"))
            vat = _to_decimal(cell(row, "vat_amount"))
            doc_type = _normalize_doc_type(cell(row, "doc_type"))
            direction = _infer_direction(doc_type, str(cell(row, "direction") or "") or None) or sheet_dir
            treatment = (
                _infer_treatment(str(cell(row, "treatment") or "") or None, rate, vat, doc_type)
                or sheet_treat
            )
            em_raw = _norm(str(cell(row, "emirate") or ""))
            emirate = _EMIRATES.get(em_raw, Emirate.UNALLOCATED)
            # Fall back to the chosen default emirate only when the row didn't specify one.
            if emirate == Emirate.UNALLOCATED and default_emirate is not None:
                emirate = default_emirate
            taxable = _to_decimal(cell(row, "taxable_amount"))
            invoice_no = _clean_id(cell(row, "invoice_number"))
            party = str(cell(row, "party") or "") or None

            # Credit notes reduce the VAT figures they relate to: negate the amounts so
            # the box aggregation SUBTRACTS them (a sales credit note lowers output VAT in
            # Box 1; a purchase/supplier credit note lowers input VAT in Box 9). Debit
            # notes keep their normal (positive) sign.
            _ttype = _norm(f"{doc_type or ''} {cell(row, 'treatment') or ''}")
            if any(k in _ttype for k in ("creditnote", "creditmemo", "salesreturn", "salesreturns")):
                if taxable is not None:
                    taxable = -abs(taxable)
                if vat is not None:
                    vat = -abs(vat)

            # Skip totals / subtotal / empty rows.
            row_text = " ".join(str(v) for v in row.values() if v is not None).lower()
            if re.search(r"\b(grand\s+)?total\b|subtotal|\bsum\b", row_text):
                continue
            if taxable == 0 and vat == 0:
                continue
            if has_id_cols and not invoice_no and not party:
                continue

            txns.append(
                Transaction(
                    row_index=i + 2,
                    date=d.isoformat() if d else (str(cell(row, "date") or "") or None),
                    doc_type=doc_type or (sheet_name or None),
                    direction=direction,
                    party=str(cell(row, "party") or "") or None,
                    trn=_clean_id(cell(row, "trn")),
                    invoice_number=invoice_no,
                    emirate=emirate,
                    treatment=treatment,
                    vat_rate=rate,
                    taxable_amount=taxable,
                    vat_amount=vat,
                    raw={"_sheet": sheet_name, **{k: (str(v) if v is not None else "") for k, v in row.items()}},
                )
            )

    # Diagnostic: if the file had rows but we couldn't recognise the essential columns,
    # tell the user exactly which headers we saw and what we expect — much more useful
    # than a generic "could not process" message.
    essential = {"taxable_amount", "vat_amount", "date", "direction", "doc_type"}
    if total_rows > 0 and not (set(last_map) & essential):
        found = ", ".join(str(h) for h in all_headers[:12] if h) or "(none)"
        raise ValueError(
            "Couldn't recognise the transaction columns in this file. "
            f"Columns found: {found}. Expected a header row with at least a Date and an "
            "amount column — e.g. Date, Type (sale/purchase), Taxable Amount, VAT Amount, "
            "Emirate, TRN, Invoice Number."
        )
    return txns, last_map
