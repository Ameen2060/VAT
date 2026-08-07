"""Build the FTA Audit File (FAF) workbook from a VAT201 return.

The FAF is the FTA's VAT-audit Excel workbook. Rather than re-create its sheets,
headers and questionnaire from scratch (and risk drifting from the official
layout), we load the bundled official template and *populate* it:

  * "Required information" — registrant name, TRN, tax periods.
  * "VAT Return"          — the box summary (amount + VAT per box, net position).
  * Box detail sheets     — one transaction row per line, mapped to the box(es)
                            the VAT201 engine assigned it to.

Everything is derived from the stored return; nothing is hardcoded per company.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

FAF_TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "faf_template.xlsx"

# Thin black border applied around every populated cell in the workbook.
_THIN = Side(style="thin", color="FF000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Column width bounds (Excel character units) for auto-fit.
_MIN_COL_W = 12
_MAX_COL_W = 42
_HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
_DATA_ALIGN = Alignment(vertical="center")


def _fit_columns(ws, ncols: int, last_row: int) -> None:
    """Size each column to its content and wrap the header so nothing is clipped.

    Column width tracks the widest *data* value (headers wrap instead of forcing
    very wide columns), bounded so short columns stay tidy and long text still fits.
    """
    for col in range(1, ncols + 1):
        # widest data value in this column
        data_w = 0
        for row in range(2, last_row + 1):
            v = ws.cell(row=row, column=col).value
            if v not in (None, ""):
                data_w = max(data_w, len(str(v)))
        # longest single word in the header (so the header can wrap cleanly)
        head = ws.cell(row=1, column=col).value
        head_word = max((len(w) for w in str(head or "").split()), default=0)
        width = max(_MIN_COL_W, min(_MAX_COL_W, max(data_w, head_word) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width
    # Wrap the header row and give it room for the wrapped lines.
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 46


def _header_width(ws) -> int:
    """Number of contiguous header columns on row 1 of a box detail sheet."""
    n = 0
    for c in range(1, (ws.max_column or 1) + 1):
        if ws.cell(row=1, column=c).value in (None, ""):
            break
        n += 1
    return n


def _border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply the thin border to every cell in the given rectangular range."""
    if max_row < min_row or max_col < min_col:
        return
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = _BORDER

# ── VAT Return summary sheet: box code -> row (Amount col C, VAT col E). ────────
_SUMMARY_ROW: dict[str, int] = {
    "1a": 7, "1b": 9, "1c": 11, "1d": 13, "1e": 15, "1f": 17, "1g": 19,
    "2": 21, "3": 23, "4": 25, "5": 27, "6": 29, "7": 31, "8": 33,
    "9": 39, "10": 41, "11": 43,
}
_SUMMARY_AMOUNT_COL = 3   # C
_SUMMARY_VAT_COL = 5      # E
# Net VAT section — single value in column B.
_NET_ROW = {"12": 50, "13": 54, "14": 58}

_STANDARD_BOXES = {"1a", "1b", "1c", "1d", "1e", "1f", "1g"}

_EMIRATE_LABEL = {
    "abu_dhabi": "Abu Dhabi", "dubai": "Dubai", "sharjah": "Sharjah",
    "ajman": "Ajman", "umm_al_quwain": "Umm Al Quwain",
    "ras_al_khaimah": "Ras Al Khaimah", "fujairah": "Fujairah",
    "unallocated": "",
}


def _num(v: Any) -> float:
    """Coerce a stored string/Decimal amount to a float for Excel numeric cells."""
    if v is None or v == "":
        return 0.0
    try:
        return float(Decimal(str(v)))
    except (InvalidOperation, ValueError):
        return 0.0


def _amt(v: Any) -> float:
    """Amount for a FAF transaction cell.

    The FTA amount/VAT columns are validated as decimals >= 1, so magnitudes are
    entered as positive numbers — the Transaction Type ("02 Credit Note") carries
    the credit sign, not a negative amount.
    """
    return abs(_num(v))


_DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d",
    "%d.%m.%Y", "%d %b %Y", "%d %B %Y", "%b %d, %Y", "%B %d, %Y",
)


def _date(v: Any) -> Any:
    """Parse a stored date string into a real ``date`` (FTA date-typed cells).

    Falls back to the original value if it cannot be parsed, so no data is lost.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (date, datetime)):
        return v
    s = str(v).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        return v


def _trn(v: Any) -> Any:
    """Coerce a TRN to an integer for the FTA whole-number cells.

    UAE TRNs are 15-digit numbers; the cells validate as whole numbers. Non-numeric
    values are left blank rather than written as text (which would fail validation).
    """
    if v is None:
        return None
    s = str(v).strip().replace(" ", "")
    return int(s) if s.isdigit() else None


def _invno(v: Any) -> Any:
    """Invoice / credit-note number, capped at the FTA's 30-character limit."""
    if v is None:
        return None
    s = str(v).strip()
    return s[:30] if s else None


def _loc(emirate: str | None) -> str:
    return _EMIRATE_LABEL.get((emirate or "").strip().lower(), "")


def _desc(t: Any) -> str:
    return (getattr(t, "doc_type", None) or "").strip()


# FTA guideline: the "Transaction Type" column on every box detail sheet is a
# dropdown restricted to exactly these two values.
FAF_INVOICE = "01 Invoice"
FAF_CREDIT_NOTE = "02 Credit Note"


def _txn_type(t: Any) -> str:
    """Classify a line as an FTA Transaction Type ("01 Invoice" | "02 Credit Note").

    A credit note is identified generically: the document type mentions a credit
    note, or the line carries a negative value (the FTA convention for credits).
    """
    dt = (getattr(t, "doc_type", None) or "").lower()
    if "credit" in dt or "cr.n" in dt or "c/n" in dt or dt.strip() in {"cn", "crn"}:
        return FAF_CREDIT_NOTE
    for attr in ("taxable_amount", "vat_amount"):
        try:
            if Decimal(str(getattr(t, attr, 0) or 0)) < 0:
                return FAF_CREDIT_NOTE
        except (InvalidOperation, ValueError):
            pass
    return FAF_INVOICE


# ── Per-sheet row builders ────────────────────────────────────────────────────
# Each takes (txn, meta) and returns the row values in the sheet's column order.
# meta = {"trn", "name", "p_from", "p_to"} (the registrant + reporting period).

def _row_box1(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            _amt(t.taxable_amount), m["p_from"], m["p_to"], _amt(t.vat_amount),
            t.party, _trn(t.trn), _desc(t), ""]


def _row_oos(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            m["p_from"], m["p_to"], _amt(t.taxable_amount), _amt(t.vat_amount),
            t.party, _trn(t.trn), _desc(t), ""]


def _row_box2(t, m) -> list:
    return [_invno(t.invoice_number), m["trn"], m["name"], _date(t.date),
            m["p_from"], m["p_to"], _amt(t.taxable_amount)]


def _row_box3(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            m["p_from"], m["p_to"], _amt(t.taxable_amount), t.party, _loc(t.emirate),
            _desc(t), _amt(t.vat_amount)]


def _row_box4(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            m["p_from"], m["p_to"], _amt(t.taxable_amount), _desc(t), t.party, _trn(t.trn),
            _loc(t.emirate)]


def _row_box5(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            m["p_from"], m["p_to"], _amt(t.taxable_amount), _desc(t), t.party, _trn(t.trn)]


def _row_box6(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            _amt(t.taxable_amount), m["p_from"], m["p_to"], t.party, _loc(t.emirate),
            "", "", _desc(t), _amt(t.vat_amount)]


def _row_box7(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), _date(t.date),
            _amt(t.taxable_amount), m["p_from"], m["p_to"], t.party, _loc(t.emirate),
            "", "", "", _amt(t.vat_amount)]


def _row_box9(t, m) -> list:
    d = _date(t.date)
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number), d,
            d, m["p_from"], m["p_to"], _amt(t.taxable_amount), t.party, _trn(t.trn),
            _desc(t), _amt(t.vat_amount), _amt(t.vat_amount), ""]


def _row_box10(t, m) -> list:
    return [_txn_type(t), m["trn"], m["name"], _invno(t.invoice_number),
            _date(t.date), m["p_from"], m["p_to"], _amt(t.taxable_amount), t.party,
            _loc(t.emirate), _desc(t), _amt(t.vat_amount)]


# sheet name -> (match predicate over a txn, row builder)
_SHEETS: list[tuple[str, Callable[[Any], bool], Callable[[Any, dict], list]]] = [
    ("Box 1", lambda t: bool(set(t.boxes or []) & _STANDARD_BOXES), _row_box1),
    ("Out of Scope Supplies", lambda t: (t.treatment or "") == "out_of_scope", _row_oos),
    ("Box 2", lambda t: "2" in (t.boxes or []), _row_box2),
    ("Box 3", lambda t: "3" in (t.boxes or []), _row_box3),
    ("Box 4", lambda t: "4" in (t.boxes or []), _row_box4),
    ("Box 5", lambda t: "5" in (t.boxes or []), _row_box5),
    ("Box 6", lambda t: "6" in (t.boxes or []), _row_box6),
    ("Box 7", lambda t: "7" in (t.boxes or []), _row_box7),
    ("Box 9", lambda t: "9" in (t.boxes or []), _row_box9),
    ("Box 10", lambda t: "10" in (t.boxes or []), _row_box10),
]


def build_faf_workbook(return_json: dict, transactions: list) -> bytes:
    """Populate the official FAF template from a return snapshot + its transactions.

    `return_json` is the stored VAT201 return dict; `transactions` is a list of
    Vat201TxnRecord (or any object exposing the same attributes).
    """
    wb = load_workbook(FAF_TEMPLATE_PATH)

    name = return_json.get("company_name") or ""
    trn = return_json.get("company_trn") or ""
    p_from = return_json.get("period_start") or ""
    p_to = return_json.get("period_end") or ""
    period_label = return_json.get("period_label") or ""
    # Coerce the constants once: registrant TRN -> int, reporting period -> real dates.
    meta = {"trn": _trn(trn), "name": name, "p_from": _date(p_from), "p_to": _date(p_to)}

    # Required information — labels merge A1:B1 / A2:B2 / A3:B3; the answer areas
    # are the merged ranges C1:G1 / C2:G2 / C3:G3, so values go in column C.
    ri = wb["Required information"]
    ri["C1"] = name
    ri["C2"] = meta["trn"] if meta["trn"] is not None else trn
    ri["C3"] = period_label

    # VAT Return summary
    vr = wb["VAT Return"]
    vr["B1"] = name
    vr["B2"] = period_label
    by_box = {b.get("box"): b for b in return_json.get("boxes", [])}
    for box, row in _SUMMARY_ROW.items():
        b = by_box.get(box)
        if not b:
            continue
        vr.cell(row=row, column=_SUMMARY_AMOUNT_COL, value=_num(b.get("amount")))
        vr.cell(row=row, column=_SUMMARY_VAT_COL, value=_num(b.get("vat")))
    for box, row in _NET_ROW.items():
        b = by_box.get(box)
        if b:
            vr.cell(row=row, column=2, value=_num(b.get("vat")))

    # Box detail sheets
    for sheet_name, match, builder in _SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ncols = _header_width(ws)
        r = 2  # row 1 is the header
        for t in transactions:
            try:
                if not match(t):
                    continue
            except Exception:
                continue
            for c, val in enumerate(builder(t, meta), start=1):
                ws.cell(row=r, column=c, value=val)
            r += 1
        # Border the whole table: header row + every data row, across all columns.
        _border_range(ws, 1, r - 1, 1, ncols)
        # Size columns to their content; wrap the (often long) headers.
        _fit_columns(ws, ncols, r - 1)

    # Border the summary sheets' populated grids too.
    _border_range(ri, 1, 17, 1, 8)          # registrant block + 12-question form (A1:H17)
    _border_range(vr, 1, 2, 1, 2)           # company name / VAT return period header
    _border_range(vr, 4, 43, 1, 9)          # VAT box grid incl. section titles (A4:I43)
    _border_range(vr, 46, 58, 1, 2)         # Net VAT Due title + due/recoverable/payable

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
